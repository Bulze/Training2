import io
import json
import math
import os
import re
import html
from datetime import date, datetime, timedelta
from concurrent.futures import ThreadPoolExecutor

import requests
from flask import Flask, jsonify, render_template, request
from flask_cors import CORS
from openpyxl import load_workbook
from dotenv import load_dotenv


DATE_HEADER = "Date/Time Europe/Belgrade"
EMP_HEADER = "Employees"
SALES_HEADER = "Sales"
GROUP_HEADER = "Group"
CREATORS_HEADER = "Creators"
TIPS_HEADER = "Tips"
DM_SALES_HEADER = "Direct message sales"
DM_SENT_HEADER = "Direct messages sent"
PPV_SALES_HEADER = "PPV sales"
FANS_CHATTED_HEADER = "Fans chatted"
FANS_SPENT_HEADER = "Fans who spent money"
FAN_CVR_HEADER = "Fan CVR"
AVG_EARNINGS_HEADER = "Avg earnings per fan who spent money"
RESP_SCHED_HEADER = "Response time (based on scheduled hours)"
RESP_CLOCK_HEADER = "Response time (based on clocked hours)"
SCHED_HOURS_HEADER = "Scheduled hours"
CLOCKED_HOURS_HEADER = "Clocked hours"
SALES_PER_HOUR_HEADER = "Sales per hour"
MSGS_PER_HOUR_HEADER = "Messages sent per hour"
FANS_PER_HOUR_HEADER = "Fans chatted per hour"

CHAT_SENDER_HEADER = "Sender"
CHAT_CREATOR_HEADER = "Creator"
CHAT_FANS_MSG_HEADER = "Fans Message"
CHAT_CREATOR_MSG_HEADER = "Creator Message"
CHAT_SENT_DATE_HEADER = "Sent date"
CHAT_REPLY_HEADER = "Replay time"
CHAT_PRICE_HEADER = "Price"
CHAT_PURCHASED_HEADER = "Purchased"
CHAT_SENT_TO_HEADER = "Sent to"


app = Flask(__name__)
load_dotenv()

def normalize_cors_origins(raw):
    if not raw:
        return "*"
    cleaned = raw.strip()
    if cleaned == "*":
        return "*"
    origins = []
    for part in cleaned.split(","):
        origin = part.strip()
        if not origin:
            continue
        if "://" in origin:
            origins.append(origin)
        else:
            origins.append(f"https://{origin}")
            origins.append(f"http://{origin}")
    return origins or "*"


CORS_ORIGINS = normalize_cors_origins(os.getenv("CORS_ORIGINS", "*"))
CORS(app, resources={r"/api/*": {"origins": CORS_ORIGINS}})

GROK_API_KEY = os.getenv("GROK_API_KEY", "")
GROK_BASE_URL = os.getenv("GROK_BASE_URL", "https://api.x.ai/v1")
GROK_MODEL = os.getenv("GROK_MODEL", "grok-2-mini")
GROK_THRESHOLD = float(os.getenv("GROK_THRESHOLD", "0.7"))


def build_grok_url():
    base = GROK_BASE_URL.rstrip("/")
    if base.endswith("/chat/completions"):
        return base
    if base.endswith("/v1"):
        return f"{base}/chat/completions"
    return f"{base}/v1/chat/completions"


def parse_json_object(text):
    if not text:
        return None
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return None
    try:
        return json.loads(text[start : end + 1])
    except Exception:
        return None


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    s = str(value)
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_date_range(value):
    if isinstance(value, datetime):
        d = value.date()
        return (d, d)
    if isinstance(value, date):
        return (value, value)
    if value is None:
        return None

    s = str(value).strip()
    # Handles strings like:
    # "2025-12-28 00:00:00 - 2026-01-08 23:59:59"
    matches = re.findall(r"(\d{4}-\d{2}-\d{2})", s)
    if not matches:
        return None

    try:
        start = datetime.strptime(matches[0], "%Y-%m-%d").date()
        end = datetime.strptime(matches[-1], "%Y-%m-%d").date()
        if end < start:
            start, end = end, start
        return (start, end)
    except ValueError:
        return None


def iter_days(start, end):
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)


def parse_money(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_hours(value):
    if value is None:
        return 0.0
    s = str(value).strip()
    if s == "-" or s == "":
        return 0.0
    hours = 0.0
    m = re.search(r"(\d+)h", s)
    if m:
        hours += float(m.group(1))
    m = re.search(r"(\d+)min", s)
    if m:
        hours += float(m.group(1)) / 60.0
    return hours


def parse_minutes(value):
    if value is None:
        return None
    s = str(value).strip()
    if s == "-" or s == "":
        return None
    minutes = 0.0
    m = re.search(r"(\d+)m", s)
    if m:
        minutes += float(m.group(1))
    m = re.search(r"(\d+)s", s)
    if m:
        minutes += float(m.group(1)) / 60.0
    return minutes


def parse_chat_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    s = str(value).strip()
    for fmt in ("%b %d, %Y", "%B %d, %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def clean_text(value):
    if value is None:
        return ""
    s = html.unescape(str(value))
    s = re.sub(r"<[^>]+>", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_text(value):
    s = clean_text(value)
    s = s.lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def build_insights(stats):
    insights = []
    sales = stats["sales"]
    clocked = stats["clocked_hours"]
    scheduled = stats["scheduled_hours"]
    sales_per_hour = stats["sales_per_hour"]
    messages_per_hour = stats["messages_per_hour"]
    response_clock = stats["response_clock_avg"]

    if sales <= 0:
        insights.append("No sales in the selected period.")
    if clocked <= 0 and scheduled > 0:
        insights.append("No clocked hours recorded; check shift tracking.")
    if clocked > 0 and sales_per_hour is not None and sales_per_hour < 100:
        insights.append("Sales per hour is below $100; focus on conversion quality.")
    if messages_per_hour is not None and messages_per_hour < 10:
        insights.append("Low messages per hour; open more active conversations.")
    if response_clock is not None and response_clock > 8:
        insights.append("Slow reply time; sales drop when response time is high.")

    if not insights:
        insights.append("Stable performance in the selected period.")
    return insights


def build_ai_insights(stats, ai_enabled=True):
    if not ai_enabled or not GROK_API_KEY:
        return build_insights(stats)

    payload = {
        "model": GROK_MODEL,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You analyze worker performance based strictly on the provided numeric metrics. "
                    "Do not infer personal activities or off-platform behavior. "
                    "Return 2-4 concise bullet points as a JSON array of strings. "
                    "Focus on measurable gaps and actionable improvements."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "sales": stats["sales"],
                        "bonus": stats["bonus"],
                        "tips": stats["tips"],
                        "ppv_sales": stats["ppv_sales"],
                        "dm_sales": stats["dm_sales"],
                        "dm_sent": stats["dm_sent"],
                        "fans_chatted": stats["fans_chatted"],
                        "fans_spent": stats["fans_spent"],
                        "clocked_hours": stats["clocked_hours"],
                        "scheduled_hours": stats["scheduled_hours"],
                        "sales_per_hour": stats["sales_per_hour"],
                        "messages_per_hour": stats["messages_per_hour"],
                        "fans_per_hour": stats["fans_per_hour"],
                        "response_clock_avg": stats["response_clock_avg"],
                        "response_sched_avg": stats["response_sched_avg"],
                    }
                ),
            },
        ],
        "temperature": 0.2,
        "max_tokens": 200,
    }

    try:
        response = requests.post(
            build_grok_url(),
            headers={"Authorization": f"Bearer {GROK_API_KEY}"},
            json=payload,
            timeout=15,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        insights = json.loads(content)
        if isinstance(insights, list) and all(isinstance(x, str) for x in insights):
            return insights
    except Exception:
        return build_insights(stats)

    return build_insights(stats)


def extract_stats(file_stream, date_from=None, date_to=None):
    wb = load_workbook(file_stream, data_only=True)
    # Prefer per-day sheet when workbook includes both totals and per-day tabs.
    ws = wb.active
    for candidate in ("By time and employee", "By Time And Employee", "By time & employee"):
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            break

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    for required in (DATE_HEADER, EMP_HEADER, SALES_HEADER):
        if required not in idx:
            raise ValueError(f"Missing column: {required}")

    min_date = None
    max_date = None
    rows = []

    for r in range(2, ws.max_row + 1):
        raw_date = ws.cell(r, idx[DATE_HEADER]).value
        date_range = parse_date_range(raw_date)
        if date_range is None:
            continue
        range_start, range_end = date_range

        if min_date is None or range_start < min_date:
            min_date = range_start
        if max_date is None or range_end > max_date:
            max_date = range_end

        rows.append((r, range_start, range_end))

    if min_date is None or max_date is None:
        raise ValueError("No valid dates found in the sheet.")

    if date_from is None:
        date_from = min_date
    if date_to is None:
        date_to = max_date

    stats = {}
    per_day = {}
    per_day_bonus = {}
    shifts = {}

    for r, range_start, range_end in rows:
        if range_end < date_from or range_start > date_to:
            continue
        overlap_start = max(range_start, date_from)
        overlap_end = min(range_end, date_to)
        if overlap_end < overlap_start:
            continue
        total_days = (range_end - range_start).days + 1
        overlap_days = (overlap_end - overlap_start).days + 1
        fraction = overlap_days / total_days if total_days else 1.0

        emp = ws.cell(r, idx[EMP_HEADER]).value
        if emp is None:
            continue

        sales_val = ws.cell(r, idx[SALES_HEADER]).value
        sales_raw = parse_money(sales_val)
        sales = sales_raw * fraction

        if emp not in stats:
            stats[emp] = {
                "sales": 0.0,
                "bonus": 0.0,
                "tips": 0.0,
                "ppv_sales": 0.0,
                "dm_sales": 0.0,
                "dm_sent": 0.0,
                "fans_chatted": 0.0,
                "fans_spent": 0.0,
                "clocked_hours": 0.0,
                "scheduled_hours": 0.0,
                "sales_per_hour_vals": [],
                "messages_per_hour_vals": [],
                "fans_per_hour_vals": [],
                "response_clock_vals": [],
                "response_sched_vals": [],
            }
            per_day[emp] = {}
            per_day_bonus[emp] = {}
            shifts[emp] = []

        stats[emp]["sales"] += sales

        if TIPS_HEADER in idx:
            stats[emp]["tips"] += parse_money(ws.cell(r, idx[TIPS_HEADER]).value) * fraction
        if PPV_SALES_HEADER in idx:
            stats[emp]["ppv_sales"] += parse_money(ws.cell(r, idx[PPV_SALES_HEADER]).value) * fraction
        if DM_SALES_HEADER in idx:
            stats[emp]["dm_sales"] += parse_money(ws.cell(r, idx[DM_SALES_HEADER]).value) * fraction
        if DM_SENT_HEADER in idx:
            stats[emp]["dm_sent"] += (parse_number(ws.cell(r, idx[DM_SENT_HEADER]).value) or 0.0) * fraction
        if FANS_CHATTED_HEADER in idx:
            stats[emp]["fans_chatted"] += (parse_number(ws.cell(r, idx[FANS_CHATTED_HEADER]).value) or 0.0) * fraction
        if FANS_SPENT_HEADER in idx:
            stats[emp]["fans_spent"] += (parse_number(ws.cell(r, idx[FANS_SPENT_HEADER]).value) or 0.0) * fraction

        if CLOCKED_HOURS_HEADER in idx:
            stats[emp]["clocked_hours"] += parse_hours(ws.cell(r, idx[CLOCKED_HOURS_HEADER]).value) * fraction
        if SCHED_HOURS_HEADER in idx:
            stats[emp]["scheduled_hours"] += parse_hours(ws.cell(r, idx[SCHED_HOURS_HEADER]).value) * fraction

        if SALES_PER_HOUR_HEADER in idx:
            v = parse_money(ws.cell(r, idx[SALES_PER_HOUR_HEADER]).value)
            if v:
                stats[emp]["sales_per_hour_vals"].append(v)
        if MSGS_PER_HOUR_HEADER in idx:
            v = parse_number(ws.cell(r, idx[MSGS_PER_HOUR_HEADER]).value)
            if v:
                stats[emp]["messages_per_hour_vals"].append(v)
        if FANS_PER_HOUR_HEADER in idx:
            v = parse_number(ws.cell(r, idx[FANS_PER_HOUR_HEADER]).value)
            if v:
                stats[emp]["fans_per_hour_vals"].append(v)

        if RESP_CLOCK_HEADER in idx:
            v = parse_minutes(ws.cell(r, idx[RESP_CLOCK_HEADER]).value)
            if v is not None:
                stats[emp]["response_clock_vals"].append(v)
        if RESP_SCHED_HEADER in idx:
            v = parse_minutes(ws.cell(r, idx[RESP_SCHED_HEADER]).value)
            if v is not None:
                stats[emp]["response_sched_vals"].append(v)

        sales_per_day = sales / overlap_days if overlap_days else sales
        for day in iter_days(overlap_start, overlap_end):
            day_key = day.isoformat()
            per_day[emp].setdefault(day_key, 0.0)
            per_day[emp][day_key] += sales_per_day
            shifts[emp].append(
                {
                    "date": day_key,
                    "group": ws.cell(r, idx[GROUP_HEADER]).value if GROUP_HEADER in idx else None,
                    "creators": ws.cell(r, idx[CREATORS_HEADER]).value if CREATORS_HEADER in idx else None,
                    "sales": sales_per_day,
                    "bonus": 0.0,
                }
            )

    # Bonus is computed per employee per day based on total daily sales (crossing $500 before midnight counts that day).
    for emp, daily in per_day.items():
        total_bonus = 0.0
        for day_key, day_sales in daily.items():
            day_bonus = math.floor(day_sales / 500.0) * 15.0
            per_day_bonus[emp][day_key] = day_bonus
            total_bonus += day_bonus
        if emp in stats:
            stats[emp]["bonus"] = total_bonus

    result = []
    for emp, data in stats.items():
        avg_sales_per_hour = (
            sum(data["sales_per_hour_vals"]) / len(data["sales_per_hour_vals"])
            if data["sales_per_hour_vals"]
            else None
        )
        avg_messages_per_hour = (
            sum(data["messages_per_hour_vals"]) / len(data["messages_per_hour_vals"])
            if data["messages_per_hour_vals"]
            else None
        )
        avg_fans_per_hour = (
            sum(data["fans_per_hour_vals"]) / len(data["fans_per_hour_vals"])
            if data["fans_per_hour_vals"]
            else None
        )
        avg_resp_clock = (
            sum(data["response_clock_vals"]) / len(data["response_clock_vals"])
            if data["response_clock_vals"]
            else None
        )
        avg_resp_sched = (
            sum(data["response_sched_vals"]) / len(data["response_sched_vals"])
            if data["response_sched_vals"]
            else None
        )

        result.append(
            {
                "employee": emp,
                "sales": data["sales"],
                "bonus": data["bonus"],
                "tips": data["tips"],
                "ppv_sales": data["ppv_sales"],
                "dm_sales": data["dm_sales"],
                "dm_sent": data["dm_sent"],
                "fans_chatted": data["fans_chatted"],
                "fans_spent": data["fans_spent"],
                "clocked_hours": data["clocked_hours"],
                "scheduled_hours": data["scheduled_hours"],
                "sales_per_hour": avg_sales_per_hour,
                "messages_per_hour": avg_messages_per_hour,
                "fans_per_hour": avg_fans_per_hour,
                "response_clock_avg": avg_resp_clock,
                "response_sched_avg": avg_resp_sched,
                "daily_sales": per_day.get(emp, {}),
                "daily_bonus": per_day_bonus.get(emp, {}),
                "shifts": shifts.get(emp, []),
            }
        )

    return {
        "min_date": min_date.isoformat(),
        "max_date": max_date.isoformat(),
        "employees": result,
    }


def extract_chat_stats(file_stream, date_from=None, date_to=None):
    wb = load_workbook(file_stream, data_only=True, read_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Chat sheet is empty.")
    headers = list(header_row)
    idx = {h: i for i, h in enumerate(headers)}

    for required in (CHAT_SENDER_HEADER, CHAT_SENT_DATE_HEADER):
        if required not in idx:
            raise ValueError(f"Missing column: {required}")

    min_date = None
    max_date = None

    stats = {}
    global_baits = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date = row[idx[CHAT_SENT_DATE_HEADER]]
        row_date = parse_chat_date(raw_date)
        if row_date is None:
            continue

        if min_date is None or row_date < min_date:
            min_date = row_date
        if max_date is None or row_date > max_date:
            max_date = row_date

        if date_from and row_date < date_from:
            continue
        if date_to and row_date > date_to:
            continue

        sender = row[idx[CHAT_SENDER_HEADER]]
        if sender is None:
            continue
        sender = str(sender).strip()
        if not sender:
            continue

        creator_msg = None
        if CHAT_CREATOR_MSG_HEADER in idx:
            creator_msg = row[idx[CHAT_CREATOR_MSG_HEADER]]
        creator_msg = clean_text(creator_msg)

        price_val = row[idx[CHAT_PRICE_HEADER]] if CHAT_PRICE_HEADER in idx else None
        price = parse_money(price_val)

        purchased_val = row[idx[CHAT_PURCHASED_HEADER]] if CHAT_PURCHASED_HEADER in idx else None
        purchased = str(purchased_val).strip().lower() if purchased_val is not None else ""
        purchased_yes = purchased in ("yes", "y", "true", "1")

        reply_val = row[idx[CHAT_REPLY_HEADER]] if CHAT_REPLY_HEADER in idx else None
        reply_min = parse_minutes(reply_val)

        sent_to = row[idx[CHAT_SENT_TO_HEADER]] if CHAT_SENT_TO_HEADER in idx else None
        sent_to = str(sent_to).strip() if sent_to is not None else ""

        if sender not in stats:
            stats[sender] = {
                "messages_sent": 0,
                "paid_offers": 0,
                "purchased": 0,
                "purchase_revenue": 0.0,
                "reply_time_vals": [],
                "unique_fans": set(),
                "sentences": {},
                "baits": {},
            }

        data = stats[sender]

        if creator_msg:
            data["messages_sent"] += 1
            key = normalize_text(creator_msg)
            if key:
                if key not in data["sentences"]:
                    data["sentences"][key] = {"text": creator_msg, "count": 0}
                data["sentences"][key]["count"] += 1

        if price > 0:
            data["paid_offers"] += 1
            if creator_msg:
                key = normalize_text(creator_msg)
                if key:
                    if key not in data["baits"]:
                        data["baits"][key] = {"text": creator_msg, "count": 0}
                    data["baits"][key]["count"] += 1
                    if key not in global_baits:
                        global_baits[key] = {"text": creator_msg, "count": 0, "purchased": 0}
                    global_baits[key]["count"] += 1

        if purchased_yes:
            data["purchased"] += 1
            data["purchase_revenue"] += price
            if creator_msg:
                key = normalize_text(creator_msg)
                if key:
                    if key not in data["baits"]:
                        data["baits"][key] = {"text": creator_msg, "count": 0}
                    data["baits"][key]["count"] += 1
                    if key not in global_baits:
                        global_baits[key] = {"text": creator_msg, "count": 0, "purchased": 0}
                    global_baits[key]["purchased"] += 1

        if reply_min is not None:
            data["reply_time_vals"].append(reply_min)

        if sent_to:
            data["unique_fans"].add(sent_to)

    if min_date is None or max_date is None:
        raise ValueError("No valid dates found in chat sheet.")

    chatters = []
    for sender, data in stats.items():
        avg_reply = (
            sum(data["reply_time_vals"]) / len(data["reply_time_vals"])
            if data["reply_time_vals"]
            else None
        )
        top_sentences = sorted(
            data["sentences"].values(), key=lambda x: x["count"], reverse=True
        )[:8]
        top_baits = sorted(data["baits"].values(), key=lambda x: x["count"], reverse=True)[:8]
        conversion = (
            data["purchased"] / data["paid_offers"] if data["paid_offers"] > 0 else None
        )

        chatters.append(
            {
                "sender": sender,
                "messages_sent": data["messages_sent"],
                "paid_offers": data["paid_offers"],
                "purchased": data["purchased"],
                "purchase_revenue": data["purchase_revenue"],
                "reply_time_avg": avg_reply,
                "unique_fans": len(data["unique_fans"]),
                "conversion_rate": conversion,
                "top_sentences": top_sentences,
                "top_baits": top_baits,
            }
        )

    return {
        "min_date": min_date.isoformat(),
        "max_date": max_date.isoformat(),
        "chatters": chatters,
        "global_baits": global_baits,
    }


def build_ppv_month(global_baits):
    items = list(global_baits.values())
    items.sort(key=lambda x: (x["purchased"], x["count"]), reverse=True)

    def classify_ppv(text):
        t = normalize_text(text)
        ass_words = ("ass", "booty", "butt", "rear", "backside")
        tits_words = ("tits", "boobs", "breasts", "titties", "chest")
        if any(w in t for w in ass_words):
            return "ass"
        if any(w in t for w in tits_words):
            return "tits"
        return "other"

    def pick(category, limit=5):
        results = []
        for item in items:
            if category == "overall" or category == classify_ppv(item["text"]):
                results.append(item)
            if len(results) >= limit:
                break
        return results

    return {
        "ass": pick("ass"),
        "tits": pick("tits"),
        "overall": pick("overall"),
    }


def build_peer_compare(employees):
    def pct(values, value, higher_better=True):
        values = [v for v in values if v is not None]
        if not values or value is None:
            return None
        total = len(values)
        if total == 1:
            return {"rank": 1, "total": 1, "percentile": 100.0}
        if higher_better:
            rank = 1 + sum(1 for v in values if v > value)
        else:
            rank = 1 + sum(1 for v in values if v < value)
        percentile = max(0.0, min(100.0, (1 - (rank - 1) / (total - 1)) * 100.0))
        return {"rank": rank, "total": total, "percentile": percentile}

    def chat_value(emp, key):
        chat = emp.get("chat") or {}
        return chat.get(key)

    metrics = {
        "sales": (lambda e: e.get("sales"), True),
        "sales_per_hour": (lambda e: e.get("sales_per_hour"), True),
        "messages_per_hour": (lambda e: e.get("messages_per_hour"), True),
        "fans_per_hour": (lambda e: e.get("fans_per_hour"), True),
        "response_clock_avg": (lambda e: e.get("response_clock_avg"), False),
        "chat_paid_offers": (lambda e: chat_value(e, "paid_offers"), True),
        "chat_conversion_rate": (lambda e: chat_value(e, "conversion_rate"), True),
        "chat_reply_time_avg": (lambda e: chat_value(e, "reply_time_avg"), False),
    }

    values_map = {}
    for key, (getter, _) in metrics.items():
        values_map[key] = [getter(e) for e in employees]

    for emp in employees:
        compare = {}
        for key, (getter, higher_better) in metrics.items():
            compare[key] = pct(values_map[key], getter(emp), higher_better)
        emp["compare"] = compare


def build_ai_chatter_summary(emp, ai_enabled=True):
    chat = emp.get("chat")
    if not chat:
        return None
    if not ai_enabled or not GROK_API_KEY:
        return {
            "why_money": ["AI is off; showing metrics and top messages."],
            "how_money": ["Use the top messages and PPV baits listed below."],
            "ppv_suggestions": ["Increase PPV volume and test the top 2 baits."],
            "bait_suggestions": ["Align copy with baits that convert."],
        }

    payload = {
        "model": GROK_MODEL,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You analyze chatter performance based strictly on the provided metrics and text. "
                    "Do not infer personal activities or off-platform behavior. "
                    "Return a JSON object with keys: why_money (array), how_money (array), "
                    "ppv_suggestions (array), bait_suggestions (array). "
                    "Each array should have 2-4 concise bullet strings."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "sales": emp.get("sales"),
                        "ppv_sales": emp.get("ppv_sales"),
                        "dm_sales": emp.get("dm_sales"),
                        "sales_per_hour": emp.get("sales_per_hour"),
                        "messages_per_hour": emp.get("messages_per_hour"),
                        "fans_per_hour": emp.get("fans_per_hour"),
                        "response_clock_avg": emp.get("response_clock_avg"),
                        "chat_messages_sent": chat.get("messages_sent"),
                        "chat_paid_offers": chat.get("paid_offers"),
                        "chat_purchased": chat.get("purchased"),
                        "chat_conversion_rate": chat.get("conversion_rate"),
                        "chat_reply_time_avg": chat.get("reply_time_avg"),
                        "top_sentences": [x["text"] for x in chat.get("top_sentences", [])],
                        "top_baits": [x["text"] for x in chat.get("top_baits", [])],
                        "compare": emp.get("compare"),
                    }
                ),
            },
        ],
        "temperature": 0.2,
        "max_tokens": 250,
    }

    try:
        response = requests.post(
            build_grok_url(),
            headers={"Authorization": f"Bearer {GROK_API_KEY}"},
            json=payload,
            timeout=15,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        summary = json.loads(content)
        if (
            isinstance(summary, dict)
            and isinstance(summary.get("why_money"), list)
            and isinstance(summary.get("how_money"), list)
            and isinstance(summary.get("ppv_suggestions"), list)
            and isinstance(summary.get("bait_suggestions"), list)
        ):
            return summary
    except Exception:
        return {
            "why_money": ["AI is unavailable; use the metrics and top messages below."],
            "how_money": ["Compare PPV/DM performance and use the best bait copy."],
            "ppv_suggestions": ["Test 2-3 PPV offers with stronger CTAs."],
            "bait_suggestions": ["Reuse the top bait messages from the list."],
        }

    return {
        "why_money": ["AI is unavailable; use the metrics and top messages below."],
        "how_money": ["Compare PPV/DM performance and use the best bait copy."],
        "ppv_suggestions": ["Test 2-3 PPV offers with stronger CTAs."],
        "bait_suggestions": ["Reuse the top bait messages from the list."],
    }


def build_ai_pair_summary(user_a, user_b, ai_enabled=True):
    if not ai_enabled or not GROK_API_KEY:
        return {
            "why_a_wins": ["Compare sales, PPV output, and reply time."],
            "why_b_lags": ["B has a lower CVR or slower reply time."],
            "ppv_recommendations": ["B should test A's PPV baits and increase paid offers."],
            "bait_recommendations": ["B should use A's top baits and iterate the copy."],
        }

    payload = {
        "model": GROK_MODEL,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Compare two chatters using only the provided metrics and text. "
                    "Return JSON with keys: why_a_wins (array), why_b_lags (array), "
                    "ppv_recommendations (array), bait_recommendations (array). "
                    "Each array should have 2-4 concise bullet strings. "
                    "Focus on measurable gaps and actionable PPV/bait improvements for user B."
                ),
            },
            {
                "role": "user",
                "content": json.dumps({"user_a": user_a, "user_b": user_b}),
            },
        ],
        "temperature": 0.2,
        "max_tokens": 250,
    }

    try:
        response = requests.post(
            build_grok_url(),
            headers={"Authorization": f"Bearer {GROK_API_KEY}"},
            json=payload,
            timeout=15,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        summary = json.loads(content)
        if (
            isinstance(summary, dict)
            and isinstance(summary.get("why_a_wins"), list)
            and isinstance(summary.get("why_b_lags"), list)
            and isinstance(summary.get("ppv_recommendations"), list)
            and isinstance(summary.get("bait_recommendations"), list)
        ):
            return summary
    except Exception:
        return {
            "why_a_wins": ["A has stronger sales/hr and better reply time."],
            "why_b_lags": ["B has lower CVR or fewer paid offers."],
            "ppv_recommendations": ["B should test A's PPV baits and tier pricing for top fans."],
            "bait_recommendations": ["B should copy A's top 2 baits and adjust tone."],
        }

    return {
        "why_a_wins": ["A has stronger sales/hr and better reply time."],
        "why_b_lags": ["B has lower CVR or fewer paid offers."],
        "ppv_recommendations": ["B should test A's PPV baits and tier pricing for top fans."],
        "bait_recommendations": ["B should copy A's top 2 baits and adjust tone."],
    }


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/analyze", methods=["POST"])
def analyze():
    try:
        if "file" not in request.files:
            return jsonify({"error": "Missing file"}), 400
        file = request.files["file"]
        if not file.filename:
            return jsonify({"error": "Empty filename"}), 400
        chat_file = request.files.get("chat_file")

        date_from = request.form.get("date_from")
        date_to = request.form.get("date_to")
        ai_enabled = request.form.get("ai_enabled", "false").lower() == "true"

        df = None
        dt = None
        if date_from:
            df = datetime.strptime(date_from, "%Y-%m-%d").date()
        if date_to:
            dt = datetime.strptime(date_to, "%Y-%m-%d").date()

        data = extract_stats(io.BytesIO(file.read()), df, dt)
        if not data["employees"]:
            return jsonify({"error": "No rows found for selected date range."}), 400

        if chat_file and chat_file.filename:
            chat_data = extract_chat_stats(io.BytesIO(chat_file.read()), df, dt)
            chat_by_sender = {c["sender"]: c for c in chat_data["chatters"]}
            data["ppv_day"] = build_ppv_month(chat_data.get("global_baits", {}))
            for emp in data["employees"]:
                emp["chat"] = chat_by_sender.get(emp["employee"])
        else:
            for emp in data["employees"]:
                emp["chat"] = None
            data["ppv_day"] = {"ass": [], "tits": [], "overall": []}

        build_peer_compare(data["employees"])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400

    ai_status = "disabled"
    if ai_enabled and not GROK_API_KEY:
        ai_status = "no_key"
    if ai_enabled and GROK_API_KEY:
        ai_status = "enabled"

    # Keep /api/analyze fast and reliable: compute non-AI insights here.
    # AI coaching is fetched per chatter via /api/employee-feedback.
    for emp in data["employees"]:
        emp["insights"] = build_insights(emp)
        emp["chat_ai"] = build_ai_chatter_summary(emp, False)

    data["ai_status"] = ai_status

    return jsonify(data)


@app.route("/api/ai-test", methods=["POST"])
def ai_test():
    if not GROK_API_KEY:
        return jsonify({"ok": False, "error": "Missing GROK_API_KEY"}), 400

    payload = {
        "model": GROK_MODEL,
        "messages": [
            {"role": "system", "content": "Reply with a JSON array: [\"ok\"]"},
            {"role": "user", "content": "ping"},
        ],
        "temperature": 0.0,
        "max_tokens": 20,
    }

    try:
        response = requests.post(
            build_grok_url(),
            headers={"Authorization": f"Bearer {GROK_API_KEY}"},
            json=payload,
            timeout=10,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        parsed = json.loads(content)
        if isinstance(parsed, list) and parsed and parsed[0] == "ok":
            return jsonify({"ok": True})
        return jsonify({"ok": False, "error": "Unexpected response"}), 400
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.route("/api/ai/evaluate", methods=["POST"])
def ai_evaluate():
    if not GROK_API_KEY:
        return jsonify({"error": "grok_not_configured"}), 503

    payload = request.get_json(silent=True) or {}
    question = payload.get("question") or ""
    ideal_answer = payload.get("idealAnswer") or payload.get("ideal_answer") or ""
    user_answer = payload.get("userAnswer") or payload.get("user_answer") or ""
    threshold = payload.get("threshold")

    if not ideal_answer or not user_answer:
        return jsonify({"error": "missing_fields"}), 400

    score_threshold = GROK_THRESHOLD
    try:
        if threshold is not None:
            score_threshold = float(threshold)
    except Exception:
        score_threshold = GROK_THRESHOLD

    prompt = f"""Evaluate semantic similarity between the ideal answer and the user answer for the question.
Return only JSON with fields: score (0 to 1) and feedback (short).
Accept paraphrases and different wording if meaning matches.

Question: {question}
Ideal answer: {ideal_answer}
User answer: {user_answer}
"""

    grok_payload = {
        "model": GROK_MODEL,
        "messages": [
            {"role": "system", "content": "Return only valid JSON. No extra text."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.2,
        "max_tokens": 200,
    }

    try:
        response = requests.post(
            build_grok_url(),
            headers={"Authorization": f"Bearer {GROK_API_KEY}"},
            json=grok_payload,
            timeout=20,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        parsed = parse_json_object(content) or {}
        score_raw = parsed.get("score")
        try:
            score = float(score_raw)
        except Exception:
            score = 0.0
        score = max(0.0, min(1.0, score))
        feedback = parsed.get("feedback") if isinstance(parsed.get("feedback"), str) else "Answer evaluated."
        correct = score >= score_threshold
        return jsonify({"correct": correct, "score": score, "feedback": feedback})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/employee-feedback", methods=["POST"])
def employee_feedback():
    if not GROK_API_KEY:
        return jsonify({"error": "Missing GROK_API_KEY"}), 400

    payload = request.get_json(silent=True) or {}
    employee = payload.get("employee") or payload.get("user") or payload.get("chatter")
    if not employee:
        return jsonify({"error": "Missing employee"}), 400

    prompt = f"""You are a performance coach for an OnlyFans chatter team.
Given the employee stats below, produce coaching feedback.
Return ONLY valid JSON with keys:
- strengths: string[] (max 4)
- improvements: string[] (max 4)
- next_steps: string[] (max 5)

Employee stats JSON:
{json.dumps(employee, ensure_ascii=False)}
"""

    grok_payload = {
        "model": GROK_MODEL,
        "messages": [
            {"role": "system", "content": "Return only valid JSON. No extra text."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.2,
        "max_tokens": 350,
    }

    try:
        response = requests.post(
            build_grok_url(),
            headers={"Authorization": f"Bearer {GROK_API_KEY}"},
            json=grok_payload,
            timeout=25,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        parsed = parse_json_object(content) or {}
        strengths = parsed.get("strengths") if isinstance(parsed.get("strengths"), list) else []
        improvements = parsed.get("improvements") if isinstance(parsed.get("improvements"), list) else []
        next_steps = parsed.get("next_steps") if isinstance(parsed.get("next_steps"), list) else []
        return jsonify({"strengths": strengths, "improvements": improvements, "next_steps": next_steps})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/compare", methods=["POST"])
def compare_users():
    try:
        payload = request.get_json(silent=True) or {}
        user_a = payload.get("user_a")
        user_b = payload.get("user_b")
        ai_enabled = bool(payload.get("ai_enabled"))
        if not user_a or not user_b:
            return jsonify({"error": "Missing users"}), 400
        summary = build_ai_pair_summary(user_a, user_b, ai_enabled)
        return jsonify({"summary": summary})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "").lower() in ("1", "true", "yes")
    app.run(host="0.0.0.0", port=port, debug=debug)
