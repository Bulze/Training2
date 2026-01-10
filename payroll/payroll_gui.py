import math
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, date

from openpyxl import load_workbook
from tkcalendar import DateEntry


DATE_HEADER = "Date/Time Europe/Belgrade"
EMP_HEADER = "Employees"
SALES_HEADER = "Sales"


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    s = str(value)
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_money(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_percent(value):
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if v > 1:
        return v / 100.0
    return v


class PayrollApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Payroll Studio")
        self.geometry("1180x720")
        self.minsize(1100, 660)

        self.file_path = tk.StringVar()
        self.date_from = tk.StringVar()
        self.date_to = tk.StringVar()

        self.default_percent = tk.StringVar(value="10")
        self.selected_employee = tk.StringVar()
        self.emp_percent = tk.StringVar()
        self.emp_penalty = tk.StringVar(value="0")
        self.status_text = tk.StringVar(value="Ready")

        self.employee_settings = {}  # emp -> {"percent": 0.1, "penalty": 0.0}
        self.employee_stats = {}  # emp -> {"sales": 0.0, "bonus": 0.0}

        self.summary_sales = tk.StringVar(value="$0.00")
        self.summary_bonus = tk.StringVar(value="$0.00")
        self.summary_total = tk.StringVar(value="$0.00")

        self._apply_styles()
        self._build_ui()

    def _apply_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        bg = "#0f172a"
        panel = "#111827"
        card = "#0b1220"
        border = "#1f2a44"
        accent = "#22d3ee"
        text = "#e5e7eb"
        muted = "#94a3b8"

        self.configure(bg=bg)

        style.configure("App.TFrame", background=bg)
        style.configure("Panel.TFrame", background=panel)
        style.configure("Card.TFrame", background=card, borderwidth=1, relief="solid")
        style.configure("CardTitle.TLabel", background=card, foreground=muted, font=("Segoe UI", 10, "bold"))
        style.configure("CardValue.TLabel", background=card, foreground=text, font=("Segoe UI", 16, "bold"))
        style.configure("Header.TFrame", background=bg)
        style.configure("HeaderTitle.TLabel", background=bg, foreground=text, font=("Segoe UI Semibold", 20))
        style.configure("HeaderSubtitle.TLabel", background=bg, foreground=muted, font=("Segoe UI", 10))

        style.configure("PanelLabel.TLabel", background=panel, foreground=muted, font=("Segoe UI", 9))
        style.configure("PanelValue.TEntry", fieldbackground="#0b1325", foreground=text, bordercolor=border)
        style.configure("Panel.TButton", background=accent, foreground="#001018", font=("Segoe UI Semibold", 10))
        style.map("Panel.TButton", background=[("active", "#38e5ff")])

        style.configure("Treeview", background=panel, fieldbackground=panel, foreground=text, rowheight=28, borderwidth=0, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", background=card, foreground=muted, font=("Segoe UI Semibold", 10), relief="flat")
        style.map("Treeview.Heading", background=[("active", card)])
        style.configure("Status.TLabel", background=bg, foreground=muted, font=("Segoe UI", 9))

    def _build_ui(self):
        container = ttk.Frame(self, style="App.TFrame")
        container.pack(fill="both", expand=True)

        header = ttk.Frame(container, style="Header.TFrame")
        header.pack(fill="x", padx=24, pady=(20, 8))
        ttk.Label(header, text="Payroll Studio", style="HeaderTitle.TLabel").pack(anchor="w")
        ttk.Label(header, text="Upload Excel, set a date range, and calculate earnings in one place.", style="HeaderSubtitle.TLabel").pack(anchor="w", pady=(4, 0))

        controls = ttk.Frame(container, style="Panel.TFrame")
        controls.pack(fill="x", padx=24, pady=(8, 12))
        controls.columnconfigure(1, weight=1)

        ttk.Label(controls, text="Excel file", style="PanelLabel.TLabel").grid(row=0, column=0, sticky="w", padx=12, pady=(12, 2))
        ttk.Entry(controls, textvariable=self.file_path, style="PanelValue.TEntry", width=70).grid(row=1, column=0, columnspan=3, sticky="ew", padx=12)
        ttk.Button(controls, text="Browse", style="Panel.TButton", command=self.browse_file).grid(row=1, column=3, padx=12, sticky="e")

        ttk.Label(controls, text="Date from", style="PanelLabel.TLabel").grid(row=2, column=0, sticky="w", padx=12, pady=(12, 2))
        self.date_from_picker = DateEntry(
            controls,
            textvariable=self.date_from,
            width=16,
            background="#0b1325",
            foreground="#e5e7eb",
            borderwidth=1,
            date_pattern="yyyy-mm-dd",
            state="readonly",
        )
        self.date_from_picker.grid(row=3, column=0, sticky="w", padx=12, pady=(0, 12))

        ttk.Label(controls, text="Date to", style="PanelLabel.TLabel").grid(row=2, column=1, sticky="w", padx=12, pady=(12, 2))
        self.date_to_picker = DateEntry(
            controls,
            textvariable=self.date_to,
            width=16,
            background="#0b1325",
            foreground="#e5e7eb",
            borderwidth=1,
            date_pattern="yyyy-mm-dd",
            state="readonly",
        )
        self.date_to_picker.grid(row=3, column=1, sticky="w", padx=12, pady=(0, 12))

        ttk.Label(controls, text="Default %", style="PanelLabel.TLabel").grid(row=2, column=2, sticky="w", padx=12, pady=(12, 2))
        ttk.Entry(controls, textvariable=self.default_percent, style="PanelValue.TEntry", width=10).grid(row=3, column=2, sticky="w", padx=12, pady=(0, 12))

        ttk.Button(controls, text="Load + Calculate", style="Panel.TButton", command=self.load_and_calculate).grid(row=3, column=3, padx=12, sticky="e")

        summary = ttk.Frame(container, style="App.TFrame")
        summary.pack(fill="x", padx=24, pady=(0, 12))
        for i in range(3):
            summary.columnconfigure(i, weight=1)

        card_sales = ttk.Frame(summary, style="Card.TFrame")
        card_sales.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        ttk.Label(card_sales, text="Total sales", style="CardTitle.TLabel").pack(anchor="w", padx=14, pady=(10, 0))
        ttk.Label(card_sales, textvariable=self.summary_sales, style="CardValue.TLabel").pack(anchor="w", padx=14, pady=(2, 12))

        card_bonus = ttk.Frame(summary, style="Card.TFrame")
        card_bonus.grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Label(card_bonus, text="Total bonus", style="CardTitle.TLabel").pack(anchor="w", padx=14, pady=(10, 0))
        ttk.Label(card_bonus, textvariable=self.summary_bonus, style="CardValue.TLabel").pack(anchor="w", padx=14, pady=(2, 12))

        card_total = ttk.Frame(summary, style="Card.TFrame")
        card_total.grid(row=0, column=2, sticky="ew", padx=(10, 0))
        ttk.Label(card_total, text="Total payout", style="CardTitle.TLabel").pack(anchor="w", padx=14, pady=(10, 0))
        ttk.Label(card_total, textvariable=self.summary_total, style="CardValue.TLabel").pack(anchor="w", padx=14, pady=(2, 12))

        table_wrap = ttk.Frame(container, style="Panel.TFrame")
        table_wrap.pack(fill="both", expand=True, padx=24, pady=(0, 12))

        columns = ("employee", "percent", "penalty", "sales", "bonus", "base_pay", "total_pay")
        self.tree = ttk.Treeview(table_wrap, columns=columns, show="headings", height=16)
        self.tree.heading("employee", text="Employee")
        self.tree.heading("percent", text="Percent")
        self.tree.heading("penalty", text="Penalty")
        self.tree.heading("sales", text="Sales")
        self.tree.heading("bonus", text="Bonus")
        self.tree.heading("base_pay", text="Base Pay")
        self.tree.heading("total_pay", text="Total Pay")

        for col in columns:
            self.tree.column(col, width=140, anchor="e")
        self.tree.column("employee", width=220, anchor="w")

        self.tree.pack(side="left", fill="both", expand=True, padx=8, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self.on_select_employee)

        scrollbar = ttk.Scrollbar(table_wrap, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y", pady=8)
        self.tree.configure(yscrollcommand=scrollbar.set)

        edit = ttk.Frame(container, style="Panel.TFrame")
        edit.pack(fill="x", padx=24, pady=(0, 12))

        ttk.Label(edit, text="Selected employee", style="PanelLabel.TLabel").grid(row=0, column=0, sticky="w", padx=12, pady=(12, 2))
        ttk.Entry(edit, textvariable=self.selected_employee, width=24, state="readonly", style="PanelValue.TEntry").grid(row=1, column=0, sticky="w", padx=12, pady=(0, 12))

        ttk.Label(edit, text="Percent", style="PanelLabel.TLabel").grid(row=0, column=1, sticky="w", padx=12, pady=(12, 2))
        ttk.Entry(edit, textvariable=self.emp_percent, width=10, style="PanelValue.TEntry").grid(row=1, column=1, sticky="w", padx=12, pady=(0, 12))

        ttk.Label(edit, text="Penalty", style="PanelLabel.TLabel").grid(row=0, column=2, sticky="w", padx=12, pady=(12, 2))
        ttk.Entry(edit, textvariable=self.emp_penalty, width=12, style="PanelValue.TEntry").grid(row=1, column=2, sticky="w", padx=12, pady=(0, 12))

        ttk.Button(edit, text="Apply", style="Panel.TButton", command=self.apply_employee_settings).grid(row=1, column=3, padx=12, sticky="e")

        footer = ttk.Frame(container, style="App.TFrame")
        footer.pack(fill="x", padx=24, pady=(0, 16))
        ttk.Label(footer, textvariable=self.status_text, style="Status.TLabel").pack(side="left")
        ttk.Button(footer, text="Clear", style="Panel.TButton", command=self.clear_data).pack(side="right", padx=(8, 0))
        ttk.Button(footer, text="Recalculate", style="Panel.TButton", command=self.refresh_table).pack(side="right")

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")],
        )
        if path:
            self.file_path.set(path)
            try:
                min_date, max_date = self._scan_date_range(path)
                self.date_from.set(min_date.strftime("%Y-%m-%d"))
                self.date_to.set(max_date.strftime("%Y-%m-%d"))
                self.date_from_picker.config(mindate=min_date, maxdate=max_date)
                self.date_to_picker.config(mindate=min_date, maxdate=max_date)
                self.status_text.set("Dates auto-detected from the Excel file.")
            except Exception as exc:
                self.status_text.set("Ready")
                messagebox.showerror("Date detection failed", f"Could not detect dates:\n{exc}")

    def load_and_calculate(self):
        path = self.file_path.get().strip()
        if not path:
            messagebox.showerror("Missing file", "Please select an Excel file.")
            return

        try:
            date_from = datetime.strptime(self.date_from.get().strip(), "%Y-%m-%d").date()
            date_to = datetime.strptime(self.date_to.get().strip(), "%Y-%m-%d").date()
        except ValueError:
            messagebox.showerror("Invalid date", "Please pick a valid date range.")
            return

        if date_from > date_to:
            messagebox.showerror("Invalid range", "Date from must be before date to.")
            return

        default_percent = parse_percent(self.default_percent.get())
        if default_percent is None or default_percent < 0:
            messagebox.showerror("Invalid percent", "Default percent must be a number.")
            return

        try:
            self.employee_stats = self._compute_stats(path, date_from, date_to)
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to load file:\n{exc}")
            return

        for emp in self.employee_stats:
            if emp not in self.employee_settings:
                self.employee_settings[emp] = {"percent": default_percent, "penalty": 0.0}

        self.refresh_table()
        self.status_text.set("Calculated for selected date range.")

    def _scan_date_range(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers)}
        if DATE_HEADER not in idx:
            raise ValueError(f"Missing column: {DATE_HEADER}")

        min_date = None
        max_date = None
        col = idx[DATE_HEADER]
        for r in range(2, ws.max_row + 1):
            raw_date = ws.cell(r, col).value
            row_date = parse_date(raw_date)
            if row_date is None:
                continue
            if min_date is None or row_date < min_date:
                min_date = row_date
            if max_date is None or row_date > max_date:
                max_date = row_date

        if min_date is None or max_date is None:
            raise ValueError("No valid dates found in the sheet.")
        return min_date, max_date

    def _compute_stats(self, path, date_from, date_to):
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers)}
        for required in (DATE_HEADER, EMP_HEADER, SALES_HEADER):
            if required not in idx:
                raise ValueError(f"Missing column: {required}")

        stats = {}

        for r in range(2, ws.max_row + 1):
            raw_date = ws.cell(r, idx[DATE_HEADER]).value
            emp = ws.cell(r, idx[EMP_HEADER]).value
            sales_val = ws.cell(r, idx[SALES_HEADER]).value

            row_date = parse_date(raw_date)
            if row_date is None or emp is None:
                continue
            if not (date_from <= row_date <= date_to):
                continue

            sales = parse_money(sales_val)
            bonus = math.floor(sales / 500.0) * 15.0

            if emp not in stats:
                stats[emp] = {"sales": 0.0, "bonus": 0.0}
            stats[emp]["sales"] += sales
            stats[emp]["bonus"] += bonus

        return stats

    def refresh_table(self):
        self.tree.delete(*self.tree.get_children())
        total_sales = 0.0
        total_bonus = 0.0
        total_payout = 0.0

        for idx, emp in enumerate(sorted(self.employee_stats.keys())):
            stats = self.employee_stats[emp]
            settings = self.employee_settings.get(emp, {"percent": 0.1, "penalty": 0.0})
            percent = settings["percent"]
            penalty = settings["penalty"]
            base_pay = stats["sales"] * percent
            total_pay = base_pay + stats["bonus"] - penalty

            total_sales += stats["sales"]
            total_bonus += stats["bonus"]
            total_payout += total_pay

            self.tree.insert(
                "",
                "end",
                values=(
                    emp,
                    f"{percent * 100:.2f}%",
                    f"${penalty:,.2f}",
                    f"${stats['sales']:,.2f}",
                    f"${stats['bonus']:,.2f}",
                    f"${base_pay:,.2f}",
                    f"${total_pay:,.2f}",
                ),
                tags=("odd" if idx % 2 else "even",),
            )

        self.tree.tag_configure("even", background="#0f172a")
        self.tree.tag_configure("odd", background="#111827")

        self.summary_sales.set(f"${total_sales:,.2f}")
        self.summary_bonus.set(f"${total_bonus:,.2f}")
        self.summary_total.set(f"${total_payout:,.2f}")

    def on_select_employee(self, _event):
        selection = self.tree.selection()
        if not selection:
            return
        row = self.tree.item(selection[0], "values")
        if not row:
            return
        emp = row[0]
        settings = self.employee_settings.get(emp, {"percent": 0.1, "penalty": 0.0})
        self.selected_employee.set(emp)
        self.emp_percent.set(f"{settings['percent'] * 100:.2f}")
        self.emp_penalty.set(f"{settings['penalty']:.2f}")

    def apply_employee_settings(self):
        emp = self.selected_employee.get()
        if not emp:
            messagebox.showerror("No employee", "Select an employee from the table.")
            return

        percent = parse_percent(self.emp_percent.get())
        if percent is None or percent < 0:
            messagebox.showerror("Invalid percent", "Percent must be a number.")
            return

        penalty = parse_money(self.emp_penalty.get())
        if penalty < 0:
            messagebox.showerror("Invalid penalty", "Penalty must be a positive number.")
            return

        self.employee_settings[emp] = {"percent": percent, "penalty": penalty}
        self.refresh_table()

    def clear_data(self):
        self.file_path.set("")
        self.date_from.set("")
        self.date_to.set("")
        self.employee_stats = {}
        self.employee_settings = {}
        self.tree.delete(*self.tree.get_children())
        self.summary_sales.set("$0.00")
        self.summary_bonus.set("$0.00")
        self.summary_total.set("$0.00")
        self.selected_employee.set("")
        self.emp_percent.set("")
        self.emp_penalty.set("0")
        self.status_text.set("Cleared. Load a new Excel file.")


if __name__ == "__main__":
    app = PayrollApp()
    app.mainloop()
